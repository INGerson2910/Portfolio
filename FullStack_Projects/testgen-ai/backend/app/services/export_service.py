from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.api.schemas_export import ExportWorkbookRequest


class ExportService:
    SUMMARY_SHEET_NAME = "Summary"
    TEST_CASES_SHEET_NAME = "Test Cases"

    HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    TITLE_FONT = Font(size=14, bold=True)
    SECTION_FONT = Font(bold=True)
    WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")

    def build_workbook(self, data: ExportWorkbookRequest) -> bytes:
        workbook = Workbook()

        summary_sheet = workbook.active
        summary_sheet.title = self.SUMMARY_SHEET_NAME
        self._populate_summary_sheet(summary_sheet, data)

        test_cases_sheet = workbook.create_sheet(self.TEST_CASES_SHEET_NAME)
        self._populate_test_cases_sheet(test_cases_sheet, data)

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return stream.getvalue()

    def _populate_summary_sheet(self, ws, data: ExportWorkbookRequest) -> None:
        ws["A1"] = "TestGen AI - Generation Summary"
        ws["A1"].font = self.TITLE_FONT

        ws["A3"] = "Feature Name"
        ws["B3"] = data.feature_name

        ws["A4"] = "Test Plan Summary"
        ws["B4"] = data.test_plan_summary

        current_row = 6

        ws[f"A{current_row}"] = "Assumptions"
        ws[f"A{current_row}"].font = self.SECTION_FONT
        current_row += 1
        for item in data.assumptions:
            ws[f"A{current_row}"] = f"- {item}"
            current_row += 1

        current_row += 1
        ws[f"A{current_row}"] = "Coverage Areas"
        ws[f"A{current_row}"].font = self.SECTION_FONT
        current_row += 1
        for item in data.coverage_areas:
            ws[f"A{current_row}"] = f"- {item}"
            current_row += 1

        current_row += 1
        ws[f"A{current_row}"] = "Warnings"
        ws[f"A{current_row}"].font = self.SECTION_FONT
        current_row += 1
        for item in data.warnings:
            ws[f"A{current_row}"] = f"- {item}"
            current_row += 1

        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 100

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = self.WRAP_ALIGNMENT

    def _populate_test_cases_sheet(self, ws, data: ExportWorkbookRequest) -> None:
        headers = [
            "TC ID",
            "Test type",
            "Description",
            "Steps",
            "Test data format",
            "Test data",
            "Expected result",
            "Pre-conditions",
            "Automatable (yes/no)",
            "Pass",
            "Actual result",
        ]

        for col_index, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_index, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.WRAP_ALIGNMENT

        for row_index, tc in enumerate(data.test_cases, start=2):
            ws.cell(row=row_index, column=1, value=tc.tc_id)
            ws.cell(row=row_index, column=2, value=tc.test_type)
            ws.cell(row=row_index, column=3, value=tc.description)
            ws.cell(row=row_index, column=4, value=tc.steps)
            ws.cell(row=row_index, column=5, value=tc.test_data_format)
            ws.cell(row=row_index, column=6, value=tc.test_data)
            ws.cell(row=row_index, column=7, value=tc.expected_result)
            ws.cell(row=row_index, column=8, value=tc.pre_conditions)
            ws.cell(row=row_index, column=9, value=tc.automatable)
            ws.cell(row=row_index, column=10, value=tc.pass_field)
            ws.cell(row=row_index, column=11, value=tc.actual_result)

        widths = {
            1: 14,
            2: 18,
            3: 40,
            4: 60,
            5: 18,
            6: 34,
            7: 45,
            8: 35,
            9: 22,
            10: 14,
            11: 30,
        }

        for col_index, width in widths.items():
            ws.column_dimensions[get_column_letter(col_index)].width = width

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=11):
            for cell in row:
                cell.alignment = self.WRAP_ALIGNMENT

        ws.freeze_panes = "A2"